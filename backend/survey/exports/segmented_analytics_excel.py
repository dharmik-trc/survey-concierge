"""
Segmented Analytics Excel Export

Exports analytics with segments side-by-side in columns.
"""

from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_segmented_analytics_excel(
    segment_analytics: Dict[str, Dict],
    questions: List,
    survey_title: str,
    segment_order: List[str],
    all_questions: List = None,
) -> openpyxl.Workbook:
    """
    Create Excel workbook with segmented analytics.

    Each question shows analytics side-by-side for all segments + "All responses".
    Comments appear under each segment column.

    Args:
        segment_analytics: Dict mapping segment_name -> analytics dict (same format as regular analytics)
        questions: List of Question objects (ordered) - may be filtered
        survey_title: Title of the survey
        segment_order: Ordered list of segment names (first should be "All responses")
        all_questions: Optional full list of all questions (ordered) - used for correct numbering

    Returns:
        openpyxl.Workbook object
    """
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws = wb.create_sheet(title="Segmented Analytics")

    # Style definitions
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    segment_header_font = Font(bold=True, size=11)
    segment_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Calculate column widths: Question column + 2 columns per segment (count, %)
    segments = segment_order or ["All responses"] + [
        s for s in segment_analytics.keys() if s != "All responses"
    ]
    num_segments = len(segments)
    cols_per_segment = 2  # Count and Percentage

    # Cache questions for quick lookup (preserves original ordering)
    question_map = {question.id: question for question in questions}

    # Create mapping from question_id to position in all_questions list (for correct numbering)
    # Use all_questions if provided, otherwise use questions (assumes questions is the full list)
    questions_for_numbering = all_questions if all_questions is not None else questions
    question_number_map = {}
    for idx, q in enumerate(questions_for_numbering, start=1):
        question_number_map[q.id] = idx

    row_num = 1

    for question in questions:
        question_id = question.id

        # Check if any segment has analytics for this question
        has_analytics = any(question_id in segment_analytics.get(seg, {}) for seg in segments)

        if not has_analytics:
            continue

        # Use position in original all_questions list to match data export numbering
        # This maintains original question numbers even when some questions don't have analytics
        question_num = question_number_map.get(question_id, (question.order or 0) + 1)

        # Question header spanning all columns
        start_col = 1
        end_col = 1 + (num_segments * cols_per_segment)
        ws.merge_cells(
            start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col
        )
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"Q{question_num}: {question.question_text}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        row_num += 1

        # Segment headers row
        ws.cell(row=row_num, column=1, value="Option/Statistic").font = segment_header_font
        ws.cell(row=row_num, column=1).fill = segment_header_fill
        ws.cell(row=row_num, column=1).border = border
        col = 2
        for seg_name in segments:
            ws.merge_cells(start_row=row_num, start_column=col, end_row=row_num, end_column=col + 1)
            cell = ws.cell(row=row_num, column=col)
            cell.value = seg_name
            cell.font = segment_header_font
            cell.fill = segment_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            col += cols_per_segment
        row_num += 1

        # Sub-headers: Count and % for each segment
        ws.cell(row=row_num, column=1, value="").fill = segment_header_fill
        ws.cell(row=row_num, column=1).border = border
        col = 2
        for seg_name in segments:
            ws.cell(row=row_num, column=col, value="Count").font = Font(bold=True, size=10)
            ws.cell(row=row_num, column=col).fill = segment_header_fill
            ws.cell(row=row_num, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col + 1, value="%").font = Font(bold=True, size=10)
            ws.cell(row=row_num, column=col + 1).fill = segment_header_fill
            ws.cell(row=row_num, column=col + 1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=row_num, column=col + 1).border = border
            col += cols_per_segment
        row_num += 1

        # Get analytics for this question from first available segment to determine type
        analytics_type = None
        for seg in segments:
            seg_analytics = segment_analytics.get(seg, {})
            q_analytics = seg_analytics.get(question_id)
            if q_analytics:
                analytics_type = q_analytics.get("type")
                break

        if analytics_type == "choice":
            row_num = _write_segmented_choice_analytics(
                ws,
                segment_analytics,
                question_id,
                segments,
                row_num,
                cols_per_segment,
                border,
                question_map,
            )
        elif analytics_type == "form_fields_numeric":
            row_num = _write_segmented_form_fields_numeric_analytics(
                ws, segment_analytics, question_id, segments, row_num, cols_per_segment, border
            )
        elif analytics_type == "numeric":
            row_num = _write_segmented_numeric_analytics(
                ws, segment_analytics, question_id, segments, row_num, cols_per_segment, border
            )
        elif analytics_type == "grid":
            row_num = _write_segmented_grid_analytics(
                ws, segment_analytics, question_id, segments, row_num, cols_per_segment, border
            )
        else:
            # Default: just show a message
            ws.cell(row=row_num, column=1, value="Analytics not available for this question type")
            row_num += 1

        # Add spacing between questions
        row_num += 1

    # Set column widths
    ws.column_dimensions["A"].width = 30
    for i in range(2, 2 + (num_segments * cols_per_segment)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

    return wb


def _write_segmented_choice_analytics(
    ws,
    segment_analytics: Dict[str, Dict],
    question_id: int,
    segments: List[str],
    start_row: int,
    cols_per_segment: int,
    border: Border,
    question_map: Dict[int, any],
) -> int:
    """Write choice analytics side-by-side for all segments.

    Supports both legacy ('options') and current ('results') analytics payloads
    and preserves the original option ordering from the survey definition.
    """

    row_num = start_row

    def _normalize_label(raw_option) -> str:
        if raw_option is None:
            return ""
        if isinstance(raw_option, dict):
            for key in ("option", "label", "value", "text"):
                if raw_option.get(key):
                    return str(raw_option[key])
            # Fallback to full dict representation
            return str(raw_option)
        return str(raw_option)

    def _extract_choice_rows(q_analytics: Dict[str, any]) -> List[Dict[str, any]]:
        if not q_analytics or q_analytics.get("type") != "choice":
            return []
        data = q_analytics.get("results")
        if not data:
            data = q_analytics.get("options", [])
        if not isinstance(data, list):
            return []
        return data

    # Preserve original option ordering from the survey definition
    ordered_options: List[str] = []
    question = question_map.get(question_id)
    if question and question.options:
        for opt in question.options:
            label = _normalize_label(opt)
            if label and label not in ordered_options:
                ordered_options.append(label)

    # Accumulate any additional options encountered in analytics (e.g. "Other")
    for seg in segments:
        q_analytics = segment_analytics.get(seg, {}).get(question_id)
        for entry in _extract_choice_rows(q_analytics):
            label = _normalize_label(entry)
            if label and label not in ordered_options:
                ordered_options.append(label)

    # Nothing to write
    if not ordered_options:
        ws.cell(row=row_num, column=1, value="No response data").border = border
        row_num += 1
        return row_num

    # Write each option with counts/percentages per segment
    for option_label in ordered_options:
        ws.cell(row=row_num, column=1, value=option_label).border = border
        col = 2
        for seg in segments:
            seg_analytics = segment_analytics.get(seg, {})
            q_analytics = seg_analytics.get(question_id)
            count = 0
            percentage = 0.0

            if q_analytics and q_analytics.get("type") == "choice":
                for entry in _extract_choice_rows(q_analytics):
                    label = _normalize_label(entry)
                    if label == option_label:
                        if isinstance(entry, dict):
                            count = entry.get("count", 0) or 0
                            percentage = entry.get("percentage")
                            if percentage is None:
                                percentage = entry.get("percent", 0)
                            percentage = percentage or 0.0
                        break

            ws.cell(row=row_num, column=col, value=count).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            ws.cell(row=row_num, column=col + 1, value=round(float(percentage), 2)).border = border
            ws.cell(row=row_num, column=col + 1).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            col += cols_per_segment
        row_num += 1

    return row_num


def _write_segmented_form_fields_numeric_analytics(
    ws,
    segment_analytics: Dict[str, Dict],
    question_id: int,
    segments: List[str],
    start_row: int,
    cols_per_segment: int,
    border: Border,
) -> int:
    """Write form_fields numeric analytics side-by-side for all segments."""
    row_num = start_row

    # Collect all subfields across all segments
    all_subfields = set()
    for seg in segments:
        seg_analytics = segment_analytics.get(seg, {})
        q_analytics = seg_analytics.get(question_id)
        if q_analytics and q_analytics.get("type") == "form_fields_numeric":
            subfields = q_analytics.get("subfields", {})
            all_subfields.update(subfields.keys())

    # Write statistics for each subfield
    for subfield_name in sorted(all_subfields):
        # Subfield header
        ws.merge_cells(
            start_row=row_num,
            start_column=1,
            end_row=row_num,
            end_column=1 + (len(segments) * cols_per_segment),
        )
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"Subfield: {subfield_name}"
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = border
        row_num += 1

        # Write stats: Count, Average, Min, Q1, Median, Q3, Max, Sum
        stats_order = ["count", "average", "min", "q1", "median", "q3", "max", "sum"]
        for stat_name in stats_order:
            stat_label = stat_name.capitalize()
            ws.cell(row=row_num, column=1, value=stat_label).border = border
            col = 2
            for seg in segments:
                seg_analytics = segment_analytics.get(seg, {})
                q_analytics = seg_analytics.get(question_id)
                value = "N/A"
                if q_analytics and q_analytics.get("type") == "form_fields_numeric":
                    subfields = q_analytics.get("subfields", {})
                    subfield_stats = subfields.get(subfield_name, {})
                    stat_value = subfield_stats.get(stat_name)
                    if stat_value is not None and stat_value != "N/A":
                        value = (
                            round(stat_value, 2)
                            if isinstance(stat_value, (int, float))
                            else stat_value
                        )

                ws.cell(row=row_num, column=col, value=value).border = border
                ws.cell(row=row_num, column=col).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
                ws.cell(row=row_num, column=col + 1, value="").border = border  # Empty % column
                col += cols_per_segment
            row_num += 1

        row_num += 1  # Spacing between subfields

    return row_num


def _write_segmented_numeric_analytics(
    ws,
    segment_analytics: Dict[str, Dict],
    question_id: int,
    segments: List[str],
    start_row: int,
    cols_per_segment: int,
    border: Border,
) -> int:
    """Write numeric analytics side-by-side for all segments."""
    row_num = start_row

    stats_order = ["count", "average", "min", "q1", "median", "q3", "max", "sum"]
    for stat_name in stats_order:
        stat_label = stat_name.capitalize()
        ws.cell(row=row_num, column=1, value=stat_label).border = border
        col = 2
        for seg in segments:
            seg_analytics = segment_analytics.get(seg, {})
            q_analytics = seg_analytics.get(question_id)
            value = "N/A"
            if q_analytics and q_analytics.get("type") == "numeric":
                stat_value = q_analytics.get(stat_name)
                if stat_value is not None and stat_value != "N/A":
                    value = (
                        round(stat_value, 2) if isinstance(stat_value, (int, float)) else stat_value
                    )

            ws.cell(row=row_num, column=col, value=value).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            ws.cell(row=row_num, column=col + 1, value="").border = border  # Empty % column
            col += cols_per_segment
        row_num += 1

    return row_num


def _write_segmented_grid_analytics(
    ws,
    segment_analytics: Dict[str, Dict],
    question_id: int,
    segments: List[str],
    start_row: int,
    cols_per_segment: int,
    border: Border,
) -> int:
    """Write grid analytics side-by-side for all segments."""
    row_num = start_row

    # Collect all rows across all segments
    # rows is a dict: {row_name: {total_responses: int, columns: [{column: str, count: int, percentage: float}]}}
    all_rows = set()
    for seg in segments:
        seg_analytics = segment_analytics.get(seg, {})
        q_analytics = seg_analytics.get(question_id)
        if q_analytics and q_analytics.get("type") == "grid":
            rows_data = q_analytics.get("rows", {})
            if isinstance(rows_data, dict):
                all_rows.update(rows_data.keys())

    # Write each row with column counts/percentages per segment
    for row_name in sorted(all_rows):
        ws.cell(row=row_num, column=1, value=f"Row: {row_name}").font = Font(bold=True, size=10)
        ws.cell(row=row_num, column=1).border = border
        row_num += 1

        # Collect all columns for this row
        all_cols = set()
        for seg in segments:
            seg_analytics = segment_analytics.get(seg, {})
            q_analytics = seg_analytics.get(question_id)
            if q_analytics and q_analytics.get("type") == "grid":
                rows_data = q_analytics.get("rows", {})
                if isinstance(rows_data, dict):
                    row_data = rows_data.get(row_name, {})
                    columns = row_data.get("columns", [])
                    for col_data in columns:
                        if isinstance(col_data, dict):
                            all_cols.add(col_data.get("column", ""))

        for col_name in sorted(all_cols):
            ws.cell(row=row_num, column=1, value=f"  {col_name}").border = border
            col = 2
            for seg in segments:
                seg_analytics = segment_analytics.get(seg, {})
                q_analytics = seg_analytics.get(question_id)
                count = 0
                percentage = 0.0
                if q_analytics and q_analytics.get("type") == "grid":
                    rows_data = q_analytics.get("rows", {})
                    if isinstance(rows_data, dict):
                        row_data = rows_data.get(row_name, {})
                        columns = row_data.get("columns", [])
                        for col_data in columns:
                            if isinstance(col_data, dict) and col_data.get("column") == col_name:
                                count = col_data.get("count", 0)
                                percentage = col_data.get("percentage", 0.0)
                                break

                ws.cell(row=row_num, column=col, value=count).border = border
                ws.cell(row=row_num, column=col).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
                ws.cell(row=row_num, column=col + 1, value=round(percentage, 2)).border = border
                ws.cell(row=row_num, column=col + 1).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
                col += cols_per_segment
            row_num += 1

        row_num += 1  # Spacing between rows

    return row_num
