"""
Export Views

This module contains the main API endpoints for exporting survey responses.
"""

import io
import json
from datetime import datetime

import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from ..count_utils import count_completed_responses
from ..models import Question, Survey
from .analytics import calculate_analytics
from .analytics_excel import create_analytics_excel
from .data_collector import (
    analyze_question_subfields,
    collect_completed_responses_only,
    collect_session_data,
    filter_sessions_by_completion,
    merge_completed_responses,
)
from .excel_builder import create_empty_sheet, create_excel_styles, create_worksheet_with_data
from .segmentation_engine import segment_sessions
from .segmented_analytics_excel import create_segmented_analytics_excel


def _create_empty_workbook(survey, survey_id):
    """
    Create an empty workbook when there are no sessions.

    Args:
        survey: Survey instance
        survey_id: UUID of the survey

    Returns:
        HttpResponse: Excel file response with a message indicating no data
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partial Responses"

    styles = create_excel_styles()

    ws.cell(row=1, column=1, value="No partial responses found")
    ws.cell(row=1, column=1).fill = styles["header_fill"]
    ws.cell(row=1, column=1).font = styles["header_font"]
    ws.cell(row=1, column=1).alignment = styles["header_alignment"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    survey_name = survey.title.replace(" ", "_")
    filename = f"{survey_name}_{current_datetime}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def export_survey_responses(request, survey_id):
    """
    Export all survey responses as an Excel file with 3 tabs grouped by session.

    This endpoint provides a comprehensive export of all survey responses organized
    into three tabs for easy analysis:

    Tabs created:
    1. **Partial Responses** - Only incomplete sessions
    2. **Completed Responses** - Only complete sessions
    3. **All Responses** - Everything combined

    Process Steps:
    - Step 1: Collect session data from partial responses table
    - Step 2: Merge complete responses from survey responses table
    - Step 3: Handle empty case (no data)
    - Step 4: Filter sessions by completion status
    - Step 5: Analyze sub-columns needed for complex questions
    - Step 6: Create workbook with 3 tabs
    - Step 7: Return Excel file

    Features:
    - Two-row header with main questions and sub-columns
    - Sequential question numbering starting from 1 (Q1, Q2, Q3...)
    - Color coding for completion status (green=completed, yellow=partial)
    - Proper formatting with frozen headers and wrapped text
    - Sub-columns for complex questions (forms, grids, etc.)

    Args:
        request: Django request object
        survey_id: UUID of the survey

    Returns:
        HttpResponse: Excel file download or error response
    """
    try:
        # Get survey and questions (allow exports for inactive surveys as well)
        survey = get_object_or_404(Survey, id=survey_id)
        all_questions = Question.objects.filter(survey=survey).order_by("order", "created_at")

        # Step 1: Collect session data from partial responses (grouped by session_id)
        sessions, completed_session_ids = collect_session_data(survey)

        # Step 2: Merge complete responses from SurveyResponse -> QuestionResponse
        merge_completed_responses(survey, sessions, completed_session_ids)

        # Step 3: Handle empty case
        if not sessions:
            return _create_empty_workbook(survey, survey_id)

        # Step 4: Filter sessions by completion status
        partial_sessions, completed_sessions = filter_sessions_by_completion(sessions)

        # Step 5: Analyze questions to identify sub-columns and multi-select questions (using all sessions)
        question_subfields, multi_select_questions = analyze_question_subfields(
            sessions, all_questions
        )

        # Step 6: Create Excel workbook with 3 tabs
        wb = openpyxl.Workbook()
        styles = create_excel_styles()

        # Remove the default sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Tab 1: Partial Responses
        if partial_sessions:
            create_worksheet_with_data(
                wb,
                "Partial Responses",
                partial_sessions,
                all_questions,
                question_subfields,
                multi_select_questions,
                styles,
            )
        else:
            create_empty_sheet(wb, "Partial Responses", "No partial responses found")

        # Tab 2: Completed Responses
        if completed_sessions:
            create_worksheet_with_data(
                wb,
                "Completed Responses",
                completed_sessions,
                all_questions,
                question_subfields,
                multi_select_questions,
                styles,
            )
        else:
            create_empty_sheet(wb, "Completed Responses", "No completed responses found")

        # Tab 3: All Responses
        create_worksheet_with_data(
            wb,
            "All Responses",
            sessions,
            all_questions,
            question_subfields,
            multi_select_questions,
            styles,
        )

        # Set "All Responses" as the active sheet (first tab user sees)
        wb.active = wb["All Responses"]

        # Step 7: Generate and return Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        survey_name = survey.title.replace(" ", "_")
        filename = f"{survey_name}_{current_datetime}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback

        error_traceback = traceback.format_exc()
        print(f"Error exporting survey responses: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response(
            {
                "error": "Failed to export survey responses",
                "details": str(e),
                "traceback": error_traceback,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def export_analytics(request, survey_id):
    """
    Export analytics for a survey as an Excel file.

    IMPORTANT: Analytics are calculated ONLY from completed survey responses
    (SurveyResponse table), NOT from partial responses.

    Calculates and exports analytics including:
    - Choice questions: count and percentage for each option
    - Grid questions: count and percentage per row per column
    - Numeric questions: min, Q1, median, Q3, max, average, sum, count

    Args:
        request: Django request object
        survey_id: UUID of the survey

    Returns:
        HttpResponse: Excel file download or error response
    """
    try:
        # Get survey and questions (allow exports for inactive surveys as well)
        survey = get_object_or_404(Survey, id=survey_id)
        all_questions = Question.objects.filter(survey=survey).order_by("order", "created_at")

        # Collect ONLY fully completed survey responses (SurveyResponse table only)
        # Explicitly excludes any PartialSurveyResponse data
        sessions, response_metadata = collect_completed_responses_only(survey)

        # Handle empty case
        if not sessions:
            # Return empty workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analytics"
            ws.cell(row=1, column=1, value="No responses available for analytics")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            survey_name = survey.title.replace(" ", "_")
            filename = f"{survey_name}_Analytics_{current_datetime}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        # Get total number of completed responses for "Answered" and "Skipped" calculations
        total_completed_responses = count_completed_responses(survey)

        # Calculate analytics (includes answered/skipped counts and comments)
        analytics = calculate_analytics(
            sessions, all_questions, total_completed_responses, response_metadata
        )

        # Create Excel workbook
        wb = create_analytics_excel(analytics, all_questions, survey.title)

        # Generate and return Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        survey_name = survey.title.replace(" ", "_")
        filename = f"{survey_name}_Analytics_{current_datetime}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback

        error_traceback = traceback.format_exc()
        print(f"Error exporting analytics: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response(
            {
                "error": "Failed to export analytics",
                "details": str(e),
                "traceback": error_traceback,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def export_segmented_analytics(request, survey_id):
    """
    Export segmented analytics for a survey as an Excel file.

    Accepts segmentation configuration in JSON body. No hardcoded fields - fully generic.

    Request Body JSON format:
    {
        "dimensions": [
            {
                "name": "Staff Size",  # Optional dimension name
                "question_id": 22,
                "type": "numeric_range",  # or "choice_mapping"
                "ranges": {
                    "Very small": [null, 5],
                    "Small": [5, 10],
                    "Medium": [10, 20],
                    "Large": [20, 50],
                    "Very large": [50, null]
                }
            },
            {
                "name": "Location",
                "question_id": 3,
                "type": "choice_mapping",
                "mapping": {
                    "London": "London & SE",
                    "South East": "London & SE",
                    "East of England": "East of England",
                    ...
                }
            }
        ]
    }

    Returns Excel file with all segments + "All responses" side-by-side in columns.
    """
    try:
        survey = get_object_or_404(Survey, id=survey_id)
        all_questions = Question.objects.filter(survey=survey).order_by("order", "created_at")

        # Collect completed responses
        sessions, response_metadata = collect_completed_responses_only(survey)

        if not sessions:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Segmented Analytics"
            ws.cell(row=1, column=1, value="No responses available for analytics")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            survey_name = survey.title.replace(" ", "_")
            filename = f"{survey_name}_Segmented_Analytics_{current_datetime}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        # Parse segmentation config from request body
        try:
            segmentation_config = json.loads(request.body.decode("utf-8") or "{}")
        except Exception as e:
            return Response(
                {"error": "Invalid JSON in request body", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not segmentation_config.get("dimensions"):
            return Response(
                {
                    "error": "No segmentation dimensions provided",
                    "message": 'Please provide at least one dimension in the "dimensions" array',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Segment sessions
        segment_to_session_ids = segment_sessions(sessions, segmentation_config, all_questions)

        # Calculate analytics for each segment
        total_completed_responses = count_completed_responses(survey)
        segment_analytics = {}

        # Calculate analytics for "All responses" first
        all_analytics = calculate_analytics(
            sessions, all_questions, total_completed_responses, response_metadata
        )
        segment_analytics["All responses"] = all_analytics

        # Calculate analytics for each segment
        for segment_name, session_ids in segment_to_session_ids.items():
            if segment_name == "All responses":
                continue  # Already handled above

            # Filter sessions for this segment
            segment_sessions_data = {sid: sessions[sid] for sid in session_ids if sid in sessions}

            if not segment_sessions_data:
                continue

            # Calculate analytics for this segment
            segment_response_count = len(segment_sessions_data)
            segment_analytics[segment_name] = calculate_analytics(
                segment_sessions_data, all_questions, segment_response_count, response_metadata
            )

        # Determine segment order: "All responses" first, then others
        segment_order = ["All responses"] + [
            seg for seg in segment_to_session_ids.keys() if seg != "All responses"
        ]

        # Create Excel workbook with segmented analytics - pass all_questions for correct numbering
        wb = create_segmented_analytics_excel(
            segment_analytics,
            all_questions,
            survey.title,
            segment_order,
            all_questions=all_questions,
        )

        # Generate and return Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        survey_name = survey.title.replace(" ", "_")
        filename = f"{survey_name}_Segmented_Analytics_{current_datetime}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback

        error_traceback = traceback.format_exc()
        print(f"Error exporting segmented analytics: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response(
            {
                "error": "Failed to export segmented analytics",
                "details": str(e),
                "traceback": error_traceback,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def preview_segmented_analytics(request, survey_id):
    """
    Preview segmented analytics for a survey as JSON data.

    Same request format as export_segmented_analytics, but returns JSON preview
    instead of Excel file. Used for real-time preview in the UI.

    Returns:
    {
        "segments": {
            "All responses": {"count": 100, "questions": {...}},
            "Small": {"count": 25, "questions": {...}},
            ...
        },
        "segment_order": ["All responses", "Small", "Medium", ...],
        "preview_data": [
            {
                "question_id": 1,
                "question_text": "...",
                "question_type": "choice",
                "segments": {
                    "All responses": {"answered": 100, "skipped": 0, "data": {...}},
                    "Small": {"answered": 25, "skipped": 0, "data": {...}},
                    ...
                }
            },
            ...
        ]
    }
    """
    try:
        survey = get_object_or_404(Survey, id=survey_id)
        all_questions = Question.objects.filter(survey=survey).order_by("order", "created_at")

        # Collect completed responses
        sessions, response_metadata = collect_completed_responses_only(survey)

        if not sessions:
            return Response(
                {
                    "segments": {},
                    "segment_order": [],
                    "preview_data": [],
                    "message": "No responses available for preview",
                },
                status=status.HTTP_200_OK,
            )

        # Parse segmentation config from request body
        try:
            segmentation_config = json.loads(request.body.decode("utf-8") or "{}")
        except Exception as e:
            return Response(
                {"error": "Invalid JSON in request body", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not segmentation_config.get("dimensions"):
            return Response(
                {
                    "segments": {},
                    "segment_order": [],
                    "preview_data": [],
                    "message": "No dimensions configured yet",
                },
                status=status.HTTP_200_OK,
            )

        # Segment sessions
        segment_to_session_ids = segment_sessions(sessions, segmentation_config, all_questions)

        # Calculate analytics for each segment
        total_completed_responses = count_completed_responses(survey)
        segment_analytics = {}
        segment_counts = {}

        # Calculate analytics for "All responses" first
        all_analytics = calculate_analytics(
            sessions, all_questions, total_completed_responses, response_metadata
        )
        segment_analytics["All responses"] = all_analytics
        segment_counts["All responses"] = total_completed_responses

        # Calculate analytics for each segment
        for segment_name, session_ids in segment_to_session_ids.items():
            if segment_name == "All responses":
                continue  # Already handled above

            # Filter sessions for this segment
            segment_sessions_data = {sid: sessions[sid] for sid in session_ids if sid in sessions}

            if not segment_sessions_data:
                segment_counts[segment_name] = 0
                continue

            # Calculate analytics for this segment
            segment_response_count = len(segment_sessions_data)
            segment_counts[segment_name] = segment_response_count
            segment_analytics[segment_name] = calculate_analytics(
                segment_sessions_data, all_questions, segment_response_count, response_metadata
            )

        # Determine segment order: "All responses" first, then others
        segment_order = ["All responses"] + [
            seg for seg in segment_to_session_ids.keys() if seg != "All responses"
        ]

        # Build preview data - simplified version for UI display
        preview_data = []
        for question in all_questions[:10]:  # Limit to first 10 questions for preview
            question_id = question.id
            question_preview = {
                "question_id": question_id,
                "question_text": question.question_text[:100]
                + ("..." if len(question.question_text) > 100 else ""),
                "question_order": question.order,
                "question_type": question.secondary_type or question.primary_type,
                "segments": {},
            }

            for segment_name in segment_order:
                if segment_name not in segment_analytics:
                    continue

                segment_question_data = segment_analytics[segment_name].get(question_id, {})

                # Extract key metrics
                answered = segment_question_data.get("answered", 0)
                skipped = segment_question_data.get("skipped", 0)

                # Simplified data preview
                preview_segment = {
                    "answered": answered,
                    "skipped": skipped,
                    "total": answered + skipped,
                }

                # Add type-specific preview
                if segment_question_data.get("type") == "choice":
                    # Show top 3 options
                    options = segment_question_data.get("results") or segment_question_data.get(
                        "options", []
                    )
                    preview_segment["top_options"] = options[:3]
                elif segment_question_data.get("type") == "numeric":
                    stats = segment_question_data.get("statistics", {})
                    preview_segment["avg"] = stats.get("mean")
                    preview_segment["min"] = stats.get("min")
                    preview_segment["max"] = stats.get("max")
                elif segment_question_data.get("type") == "form_fields_numeric":
                    subfields = segment_question_data.get("subfields", {})
                    preview_segment["subfield_count"] = len(subfields)

                question_preview["segments"][segment_name] = preview_segment

            preview_data.append(question_preview)

        # Build response
        segments_info = {}
        for seg_name in segment_order:
            segments_info[seg_name] = {"count": segment_counts.get(seg_name, 0)}

        return Response(
            {
                "segments": segments_info,
                "segment_order": segment_order,
                "preview_data": preview_data,
                "total_questions": len(all_questions),
                "showing_questions": len(preview_data),
            },
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        import traceback

        error_traceback = traceback.format_exc()
        print(f"Error previewing segmented analytics: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response(
            {"error": "Failed to preview segmented analytics", "details": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def export_filtered_analytics(request, survey_id):
    """
    Export filtered analytics for a survey as an Excel file.

    Filters responses to only those matching selected options from a choice question.
    Uses OR logic - if response matches ANY of the selected options, it's included.

    Request Body JSON format:
    {
        "question_id": 3,
        "selected_options": ["London", "South East", "East Midlands"]
    }

    Returns Excel file with analytics for filtered responses only.
    """
    try:
        survey = get_object_or_404(Survey, id=survey_id)
        all_questions = Question.objects.filter(survey=survey).order_by("order", "created_at")

        # Collect completed responses
        sessions, response_metadata = collect_completed_responses_only(survey)

        if not sessions:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Filtered Analytics"
            ws.cell(row=1, column=1, value="No responses available for analytics")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            survey_name = survey.title.replace(" ", "_")
            filename = f"{survey_name}_Filtered_Analytics_{current_datetime}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        # Parse filter config from request body
        try:
            request_body_str = request.body.decode("utf-8") or "{}"
            filter_config = json.loads(request_body_str)
        except Exception as e:
            return Response(
                {"error": "Invalid JSON in request body", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get exclude_open_text option (default: False for backward compatibility)
        exclude_open_text = filter_config.get("exclude_open_text", False)
        # Ensure it's a boolean (handle string "true"/"false" from frontend if needed)
        if isinstance(exclude_open_text, str):
            exclude_open_text = exclude_open_text.lower() in ("true", "1", "yes")
        exclude_open_text = bool(exclude_open_text)

        # Support both old format (single filter) and new format (multiple filters)
        filters_list = filter_config.get("filters", [])
        if not filters_list:
            # Backward compatibility: check for old format
            filter_question_id = filter_config.get("question_id")
            selected_options = filter_config.get("selected_options", [])
            if filter_question_id and selected_options:
                filters_list = [
                    {"question_id": filter_question_id, "selected_options": selected_options}
                ]

        if not filters_list:
            return Response(
                {
                    "error": 'No filters provided. Use "filters" array with question_id and selected_options (for choice) or numeric_range (for numeric) for each filter.'
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Import helper function for numeric parsing
        from .segmentation_engine import _get_total_fte_from_form_fields, _parse_numeric_value

        # Validate all filters
        filter_questions = {}
        for i, filter_item in enumerate(filters_list):
            filter_question_id = filter_item.get("question_id")
            selected_options = filter_item.get("selected_options", [])
            numeric_range = filter_item.get("numeric_range")  # [min, max] or None

            if not filter_question_id:
                return Response(
                    {"error": f"Filter {i+1}: question_id is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Verify question exists
            filter_question = Question.objects.filter(id=filter_question_id, survey=survey).first()
            if not filter_question:
                return Response(
                    {
                        "error": f"Filter {i+1}: Question {filter_question_id} not found for this survey"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Determine filter type: choice or numeric_range
            if numeric_range is not None:
                # Numeric range filter
                if not isinstance(numeric_range, list) or len(numeric_range) != 2:
                    return Response(
                        {
                            "error": f"Filter {i+1}: numeric_range must be [min, max] array (use null for unbounded)"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Validate question is numeric
                is_numeric = (
                    filter_question.primary_type == "open_text"
                    and filter_question.secondary_type
                    in ["number", "positive_number", "negative_number"]
                ) or (
                    filter_question.primary_type == "form"
                    and filter_question.secondary_type == "form_fields"
                )

                if not is_numeric:
                    return Response(
                        {
                            "error": f"Filter {i+1}: numeric_range filtering is only available for numeric questions"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                filter_questions[filter_question_id] = {
                    "question": filter_question,
                    "type": "numeric_range",
                    "numeric_range": numeric_range,
                }
            elif selected_options and isinstance(selected_options, list):
                # Choice filter
                if (
                    filter_question.primary_type != "form"
                    or filter_question.secondary_type
                    not in ["radio", "dropdown", "multiple_choices"]
                ):
                    return Response(
                        {
                            "error": f"Filter {i+1}: Filtering is only available for choice questions (radio, dropdown, multiple_choices)"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                filter_questions[filter_question_id] = {
                    "question": filter_question,
                    "type": "choice",
                    "selected_options": selected_options,
                }
            else:
                return Response(
                    {
                        "error": f"Filter {i+1}: Must provide either selected_options (for choice) or numeric_range (for numeric)"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Filter sessions - responses must match ALL filters (AND logic)
        # Within each filter, responses matching ANY option are included (OR logic for choice)
        filtered_sessions = {}
        for session_id, session_data in sessions.items():
            answers = session_data.get("questions", {})
            matches_all_filters = True

            for filter_question_id, filter_info in filter_questions.items():
                answer = answers.get(filter_question_id)
                filter_type = filter_info.get("type")

                if filter_type == "choice":
                    # Choice filter: check if ANY answer value matches ANY selected option
                    answer_values = []
                    if isinstance(answer, dict) and "answer" in answer:
                        answer = answer.get("answer")

                    if isinstance(answer, list):
                        answer_values = [str(x) for x in answer]
                    elif answer is not None:
                        answer_values = [str(answer)]

                    selected_options_str = [str(opt) for opt in filter_info["selected_options"]]
                    matches_this_filter = any(
                        str(answer_val) in selected_options_str for answer_val in answer_values
                    )

                elif filter_type == "numeric_range":
                    # Numeric range filter: check if value falls within range
                    from .analytics import _is_blank_value, _is_effectively_blank_answer

                    question = filter_info["question"]
                    range_min, range_max = filter_info["numeric_range"]

                    # Skip blank answers entirely - they don't match any range
                    if _is_effectively_blank_answer(answer):
                        matches_this_filter = False
                    else:
                        # Parse numeric value
                        numeric_value = None
                        if (
                            question.primary_type == "form"
                            and question.secondary_type == "form_fields"
                        ):
                            numeric_value = _get_total_fte_from_form_fields(answer)
                        else:
                            # Single numeric value
                            if isinstance(answer, dict) and "answer" in answer:
                                inner_answer = answer.get("answer")
                                # Skip if inner answer is blank
                                if _is_blank_value(inner_answer):
                                    numeric_value = None
                                else:
                                    numeric_value = _parse_numeric_value(inner_answer)
                            else:
                                numeric_value = _parse_numeric_value(answer)

                        if numeric_value is None:
                            matches_this_filter = False
                        else:
                            # Check if value is within range
                            matches_min = range_min is None or numeric_value >= float(range_min)
                            matches_max = range_max is None or numeric_value <= float(range_max)
                            matches_this_filter = matches_min and matches_max
                else:
                    matches_this_filter = False

                if not matches_this_filter:
                    matches_all_filters = False
                    break  # No need to check other filters if this one doesn't match

            if matches_all_filters:
                filtered_sessions[session_id] = session_data

        if not filtered_sessions:
            # Build filter description for error message
            filter_descriptions = []
            for filter_question_id, filter_info in filter_questions.items():
                q = filter_info["question"]
                filter_type = filter_info.get("type", "choice")
                if filter_type == "numeric_range":
                    range_min, range_max = filter_info.get("numeric_range", [None, None])
                    range_desc = f"[{range_min if range_min is not None else 'min'}, {range_max if range_max is not None else 'max'}]"
                    filter_descriptions.append(f"Q{q.order + 1} {range_desc}")
                else:
                    opts = filter_info.get("selected_options", [])
                    if opts:
                        filter_descriptions.append(
                            f"Q{q.order + 1} ({', '.join(opts[:3])}{'...' if len(opts) > 3 else ''})"
                        )
                    else:
                        filter_descriptions.append(f"Q{q.order + 1} (invalid filter)")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Filtered Analytics"
            ws.cell(
                row=1,
                column=1,
                value=f"No responses match all filters: {' AND '.join(filter_descriptions)}",
            )

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            survey_name = survey.title.replace(" ", "_")
            filename = f"{survey_name}_Filtered_Analytics_{current_datetime}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        # Filter questions if exclude_open_text is True
        questions_to_include = all_questions
        if exclude_open_text:
            questions_to_include = [
                q
                for q in all_questions
                if not (q.primary_type == "open_text" and q.secondary_type in ["text", "paragraph"])
            ]

        # Calculate analytics for filtered sessions
        filtered_response_count = len(filtered_sessions)
        analytics = calculate_analytics(
            filtered_sessions, questions_to_include, filtered_response_count, response_metadata
        )

        # Create Excel workbook - pass all_questions for correct numbering
        wb = create_analytics_excel(
            analytics, questions_to_include, survey.title, all_questions=all_questions
        )

        # Generate and return Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        survey_name = survey.title.replace(" ", "_")
        # Build filename from filters
        filter_strs = []
        for filter_question_id, filter_info in filter_questions.items():
            filter_type = filter_info.get("type", "choice")
            if filter_type == "numeric_range":
                range_min, range_max = filter_info.get("numeric_range", [None, None])
                range_label = (
                    f"{range_min if range_min is not None else 'min'}-"
                    f"{range_max if range_max is not None else 'max'}"
                )
                filter_strs.append(range_label.replace(" ", "_"))
            else:
                opts = filter_info.get("selected_options", [])
                if opts:
                    filter_str = "_".join(opts[:2]).replace(" ", "_")
                    if len(opts) > 2:
                        filter_str += f"_+{len(opts)-2}more"
                else:
                    filter_str = f"Q{filter_question_id}_choice"
                filter_strs.append(filter_str)

        if filter_strs:
            filter_str = "_".join(filter_strs[:3])
            if len(filter_strs) > 3:
                filter_str += f"_+{len(filter_strs)-3}more_filters"
            filename = f"{survey_name}_Filtered_{filter_str}_{current_datetime}.xlsx"
        else:
            filename = f"{survey_name}_Filtered_{current_datetime}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback

        error_traceback = traceback.format_exc()
        print(f"Error exporting filtered analytics: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response(
            {
                "error": "Failed to export filtered analytics",
                "details": str(e),
                "traceback": error_traceback,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([AllowAny])
def preview_filtered_analytics(request, survey_id):
    """
    Preview filtered analytics for a survey as JSON data.

    Same request format as export_filtered_analytics, but returns JSON preview
    instead of Excel file. Used for real-time preview in the UI.

    Request Body JSON format:
    {
        "filters": [
            {
                "question_id": 3,
                "selected_options": ["London", "South East"]
            },
            {
                "question_id": 5,
                "selected_options": ["Option A", "Option B"]
            }
        ]
    }

    OR old format (backward compatible):
    {
        "question_id": 3,
        "selected_options": ["London", "South East"]
    }

    Returns:
    {
        "filtered_count": 25,
        "total_count": 100,
        "filters": [
            {
                "filter_question": "Q3: What is your location?",
                "selected_options": ["London", "South East"]
            },
            ...
        ],
        "preview_data": [...]
    }
    """
    try:
        survey = get_object_or_404(Survey, id=survey_id)
        all_questions = Question.objects.filter(survey=survey).order_by("order", "created_at")

        # Collect completed responses
        sessions, response_metadata = collect_completed_responses_only(survey)

        if not sessions:
            return Response(
                {
                    "filtered_count": 0,
                    "total_count": 0,
                    "filter_question": "",
                    "selected_options": [],
                    "preview_data": [],
                    "message": "No responses available for preview",
                },
                status=status.HTTP_200_OK,
            )

        # Parse filter config from request body
        try:
            filter_config = json.loads(request.body.decode("utf-8") or "{}")
        except Exception as e:
            return Response(
                {"error": "Invalid JSON in request body", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get exclude_open_text option (default: False for backward compatibility)
        exclude_open_text = filter_config.get("exclude_open_text", False)
        # Ensure it's a boolean (handle string "true"/"false" from frontend if needed)
        if isinstance(exclude_open_text, str):
            exclude_open_text = exclude_open_text.lower() in ("true", "1", "yes")
        exclude_open_text = bool(exclude_open_text)

        # Support both old format (single filter) and new format (multiple filters)
        filters_list = filter_config.get("filters", [])
        if not filters_list:
            # Backward compatibility: check for old format
            filter_question_id = filter_config.get("question_id")
            selected_options = filter_config.get("selected_options", [])
            if filter_question_id and selected_options:
                filters_list = [
                    {"question_id": filter_question_id, "selected_options": selected_options}
                ]

        if not filters_list:
            return Response(
                {
                    "filtered_count": 0,
                    "total_count": len(sessions),
                    "filters": [],
                    "preview_data": [],
                    "message": "No filter configured yet",
                },
                status=status.HTTP_200_OK,
            )

        # Import helper function for numeric parsing
        from .segmentation_engine import _get_total_fte_from_form_fields, _parse_numeric_value

        # Validate all filters
        filter_questions = {}
        filter_info_list = []
        for i, filter_item in enumerate(filters_list):
            filter_question_id = filter_item.get("question_id")
            selected_options = filter_item.get("selected_options", [])
            numeric_range = filter_item.get("numeric_range")

            if not filter_question_id:
                continue  # Skip invalid filters

            # Verify question exists
            filter_question = Question.objects.filter(id=filter_question_id, survey=survey).first()
            if not filter_question:
                continue  # Skip invalid questions

            # Determine filter type
            if numeric_range is not None:
                # Numeric range filter
                if not isinstance(numeric_range, list) or len(numeric_range) != 2:
                    continue  # Skip invalid numeric range

                # Validate question is numeric
                is_numeric = (
                    filter_question.primary_type == "open_text"
                    and filter_question.secondary_type
                    in ["number", "positive_number", "negative_number"]
                ) or (
                    filter_question.primary_type == "form"
                    and filter_question.secondary_type == "form_fields"
                )

                if not is_numeric:
                    continue  # Skip non-numeric questions

                filter_questions[filter_question_id] = {
                    "question": filter_question,
                    "type": "numeric_range",
                    "numeric_range": numeric_range,
                }
                range_min, range_max = numeric_range
                range_desc = f"[{range_min if range_min is not None else 'min'}, {range_max if range_max is not None else 'max'}]"
                filter_info_list.append(
                    {
                        "filter_question": f"Q{filter_question.order + 1}: {filter_question.question_text[:100]}",
                        "numeric_range": range_desc,
                    }
                )
            elif selected_options and isinstance(selected_options, list):
                # Choice filter
                if (
                    filter_question.primary_type != "form"
                    or filter_question.secondary_type
                    not in ["radio", "dropdown", "multiple_choices"]
                ):
                    continue  # Skip non-choice questions

                filter_questions[filter_question_id] = {
                    "question": filter_question,
                    "type": "choice",
                    "selected_options": selected_options,
                }
                filter_info_list.append(
                    {
                        "filter_question": f"Q{filter_question.order + 1}: {filter_question.question_text[:100]}",
                        "selected_options": selected_options,
                    }
                )
            else:
                continue  # Skip invalid filters

        if not filter_questions:
            return Response(
                {
                    "filtered_count": 0,
                    "total_count": len(sessions),
                    "filters": [],
                    "preview_data": [],
                    "message": "No valid filters configured",
                },
                status=status.HTTP_200_OK,
            )

        # Filter sessions - responses must match ALL filters (AND logic)
        # Within each filter, responses matching ANY option are included (OR logic for choice)
        filtered_sessions = {}
        for session_id, session_data in sessions.items():
            answers = session_data.get("questions", {})
            matches_all_filters = True

            for filter_question_id, filter_info in filter_questions.items():
                answer = answers.get(filter_question_id)
                filter_type = filter_info.get("type")

                if filter_type == "choice":
                    # Choice filter: check if ANY answer value matches ANY selected option
                    answer_values = []
                    if isinstance(answer, dict) and "answer" in answer:
                        answer = answer.get("answer")

                    if isinstance(answer, list):
                        answer_values = [str(x) for x in answer]
                    elif answer is not None:
                        answer_values = [str(answer)]

                    selected_options_str = [str(opt) for opt in filter_info["selected_options"]]
                    matches_this_filter = any(
                        str(answer_val) in selected_options_str for answer_val in answer_values
                    )

                elif filter_type == "numeric_range":
                    # Numeric range filter: check if value falls within range
                    from .analytics import _is_blank_value, _is_effectively_blank_answer

                    question = filter_info["question"]
                    range_min, range_max = filter_info["numeric_range"]

                    # Skip blank answers entirely - they don't match any range
                    if _is_effectively_blank_answer(answer):
                        matches_this_filter = False
                    else:
                        # Parse numeric value
                        numeric_value = None
                        if (
                            question.primary_type == "form"
                            and question.secondary_type == "form_fields"
                        ):
                            numeric_value = _get_total_fte_from_form_fields(answer)
                        else:
                            # Single numeric value
                            if isinstance(answer, dict) and "answer" in answer:
                                inner_answer = answer.get("answer")
                                # Skip if inner answer is blank
                                if _is_blank_value(inner_answer):
                                    numeric_value = None
                                else:
                                    numeric_value = _parse_numeric_value(inner_answer)
                            else:
                                numeric_value = _parse_numeric_value(answer)

                        if numeric_value is None:
                            matches_this_filter = False
                        else:
                            # Check if value is within range
                            matches_min = range_min is None or numeric_value >= float(range_min)
                            matches_max = range_max is None or numeric_value <= float(range_max)
                            matches_this_filter = matches_min and matches_max
                else:
                    matches_this_filter = False

                if not matches_this_filter:
                    matches_all_filters = False
                    break  # No need to check other filters if this one doesn't match

            if matches_all_filters:
                filtered_sessions[session_id] = session_data

        filtered_count = len(filtered_sessions)
        total_count = len(sessions)

        if not filtered_sessions:
            filter_descriptions = []
            for filter_info in filter_info_list:
                if "selected_options" in filter_info:
                    opts = filter_info["selected_options"]
                    filter_descriptions.append(
                        f"{filter_info['filter_question']} ({', '.join(opts[:3])}{'...' if len(opts) > 3 else ''})"
                    )
                elif "numeric_range" in filter_info:
                    filter_descriptions.append(
                        f"{filter_info['filter_question']} {filter_info['numeric_range']}"
                    )

            filter_desc_str = " AND ".join(filter_descriptions)
            return Response(
                {
                    "filtered_count": 0,
                    "total_count": total_count,
                    "filters": filter_info_list,
                    "preview_data": [],
                    "message": f"No responses match all filters: {filter_desc_str}",
                },
                status=status.HTTP_200_OK,
            )

        # Filter questions if exclude_open_text is True
        questions_to_include = all_questions
        open_text_excluded_count = 0
        if exclude_open_text:
            # Count open text questions that will be excluded
            open_text_questions = [
                q
                for q in all_questions
                if q.primary_type == "open_text" and q.secondary_type in ["text", "paragraph"]
            ]
            open_text_excluded_count = len(open_text_questions)
            questions_to_include = [
                q
                for q in all_questions
                if not (q.primary_type == "open_text" and q.secondary_type in ["text", "paragraph"])
            ]
            filtered_count = len(questions_to_include)

        # Calculate analytics for filtered sessions
        analytics = calculate_analytics(
            filtered_sessions, questions_to_include, filtered_count, response_metadata
        )

        # Build preview data - simplified version for UI display
        preview_data = []
        for question in questions_to_include[:10]:  # Limit to first 10 questions for preview
            question_id = question.id
            question_data = analytics.get(question_id, {})

            answered = question_data.get("answered", 0)
            skipped = question_data.get("skipped", 0)

            question_preview = {
                "question_id": question_id,
                "question_text": question.question_text[:100]
                + ("..." if len(question.question_text) > 100 else ""),
                "question_order": question.order,
                "question_type": question.secondary_type or question.primary_type,
                "answered": answered,
                "skipped": skipped,
                "total": answered + skipped,
            }

            # Add type-specific preview
            if question_data.get("type") == "choice":
                # Show top 3 options
                options = question_data.get("results") or question_data.get("options", [])
                question_preview["top_options"] = options[:3]
            elif question_data.get("type") == "numeric":
                stats = question_data.get("statistics", {})
                question_preview["avg"] = stats.get("mean")
                question_preview["min"] = stats.get("min")
                question_preview["max"] = stats.get("max")
            elif question_data.get("type") == "form_fields_numeric":
                subfields = question_data.get("subfields", {})
                question_preview["subfield_count"] = len(subfields)

            preview_data.append(question_preview)

        return Response(
            {
                "filtered_count": filtered_count,
                "total_count": total_count,
                "filters": filter_info_list,
                "preview_data": preview_data,
                "total_questions": len(questions_to_include),
                "showing_questions": len(preview_data),
                "exclude_open_text": exclude_open_text,
                "open_text_excluded_count": open_text_excluded_count,
                "total_questions_before_filter": (
                    len(all_questions) if exclude_open_text else len(questions_to_include)
                ),
            },
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        import traceback

        error_traceback = traceback.format_exc()
        print(f"Error previewing filtered analytics: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response(
            {"error": "Failed to preview filtered analytics", "details": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
