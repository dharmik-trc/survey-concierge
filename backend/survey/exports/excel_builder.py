"""
Excel Building Utilities

This module contains functions for building and formatting Excel workbooks.
"""

import json

import openpyxl
from openpyxl.styles import Alignment

from .excel_styles import create_excel_styles


def build_excel_headers(ws, all_questions, question_subfields, multi_select_questions, styles):
    """
    Build two-row header structure for Excel worksheet.

    Creates a professional two-row header where:
    - Row 1: Main question headers (merged across sub-columns if applicable)
    - Row 2: Sub-column headers for questions with multiple fields
    - Multi-select questions: Each option gets its own column (Q1_1, Q1_2, etc.)

    Args:
        ws: Worksheet object
        all_questions: QuerySet or list of Question objects
        question_subfields: Dict mapping question_id to list of subfield names
        multi_select_questions: Dict mapping question_id to list of options
        styles: Dict of style objects from create_excel_styles()

    Returns:
        tuple: (question_column_map, is_completed_col, last_activity_col)
            - question_column_map: Dict mapping question IDs to column indices
            - is_completed_col: Column index for "Is Completed" field
            - last_activity_col: Column index for "Last Activity" field
    """
    question_column_map = {}
    col_num = (
        4  # Start at column 4 (column 1 is #, column 2 is Is Completed, column 3 is Respondent ID)
    )
    question_num = 0

    # Number column header (single row)
    cell = ws.cell(row=1, column=1)
    cell.value = "#"
    cell.fill = styles["header_fill"]
    cell.font = styles["header_font"]
    cell.alignment = styles["header_alignment"]

    # Is Completed column header (single row)
    cell = ws.cell(row=1, column=2)
    cell.value = "Is Completed"
    cell.fill = styles["header_fill"]
    cell.font = styles["header_font"]
    cell.alignment = styles["header_alignment"]

    # Respondent ID column header (single row)
    cell = ws.cell(row=1, column=3)
    cell.value = "Respondent ID"
    cell.fill = styles["header_fill"]
    cell.font = styles["header_font"]
    cell.alignment = styles["header_alignment"]

    # Question columns
    for question in all_questions:
        question_num += 1
        question_text = question.question_text

        if question.id in multi_select_questions:
            # Multi-select question - each option gets its own column
            options = multi_select_questions[question.id]
            start_col = col_num
            end_col = col_num + len(options) - 1
            question_column_map[question.id] = {"_type": "multi_select", "_options": {}}

            # Row 1: Full question repeated in each column
            for col in range(start_col, end_col + 1):
                main_cell = ws.cell(row=1, column=col)
                main_cell.value = f"Q{question_num}: {question_text}"
                main_cell.fill = styles["header_fill"]
                main_cell.font = styles["header_font"]
                main_cell.alignment = styles["header_alignment"]

            # Row 2: Option columns with Q#_# format
            for idx, option in enumerate(options, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = f"Q{question_num}_{idx}: {option}"
                cell.fill = styles["subheader_fill"]
                cell.font = styles["subheader_font"]
                cell.alignment = styles["header_alignment"]
                question_column_map[question.id]["_options"][option] = col_num
                col_num += 1

        elif question.id in question_subfields:
            # Question with sub-columns (forms, grids, etc.)
            subfields = question_subfields[question.id]
            start_col = col_num
            end_col = col_num + len(subfields) - 1
            question_column_map[question.id] = {}

            # Row 1: Full question repeated in each column
            for col in range(start_col, end_col + 1):
                main_cell = ws.cell(row=1, column=col)
                main_cell.value = f"Q{question_num}: {question_text}"
                main_cell.fill = styles["header_fill"]
                main_cell.font = styles["header_font"]
                main_cell.alignment = styles["header_alignment"]

            # Row 2: Sub-column headers with "id: text" format
            for idx, subfield in enumerate(subfields, 1):
                subfield_display = subfield.replace("_", " ").title()
                cell = ws.cell(row=2, column=col_num)
                cell.value = f"Q{question_num}_{idx}: {subfield_display}"
                cell.fill = styles["subheader_fill"]
                cell.font = styles["subheader_font"]
                cell.alignment = styles["header_alignment"]
                question_column_map[question.id][subfield] = col_num
                col_num += 1
        else:
            # Simple question - single column with two-row header
            # Row 1: Main question text
            cell = ws.cell(row=1, column=col_num)
            cell.value = f"Q{question_num}: {question_text}"
            cell.fill = styles["header_fill"]
            cell.font = styles["header_font"]
            cell.alignment = styles["header_alignment"]

            # Row 2: Repeat question text for consistency with other question types
            sub_cell = ws.cell(row=2, column=col_num)
            sub_cell.value = f"Q{question_num}: {question_text}"
            sub_cell.fill = styles["subheader_fill"]
            sub_cell.font = styles["subheader_font"]
            sub_cell.alignment = styles["header_alignment"]

            question_column_map[question.id] = {"_main": col_num}
            col_num += 1

    # Status columns (Is Completed is already at column 2, so only Last Activity remains)
    is_completed_col = 2  # Fixed at column 2
    last_activity_col = col_num  # After all question columns

    # Last Activity column (single row)
    cell = ws.cell(row=1, column=last_activity_col)
    cell.value = "Last Activity"
    cell.fill = styles["header_fill"]
    cell.font = styles["header_font"]
    cell.alignment = styles["header_alignment"]

    return question_column_map, is_completed_col, last_activity_col


def write_answer_to_cell(ws, row_num, col_idx, value, alignment):
    """
    Write a value to a cell with proper formatting.

    Handles conversion of various data types (lists, dicts, primitives) to string format.

    Args:
        ws: Worksheet object
        row_num: Row number to write to
        col_idx: Column index to write to
        value: Value to write (can be list, dict, string, number, etc.)
        alignment: Alignment object for the cell

    Returns:
        Cell: The modified cell object
    """
    if isinstance(value, list):
        value_str = ", ".join([str(item) for item in value])
    elif isinstance(value, dict):
        # Check if this is a combined answer + comment object
        if "answer" in value and "comment" in value:
            answer_part = value["answer"]
            comment_part = value["comment"]
            # Format as "Answer - Comment" for better readability
            if comment_part and str(comment_part).strip():
                value_str = f"{answer_part} - {comment_part}"
            else:
                value_str = str(answer_part)
        else:
            # For other dict types, use JSON format
            value_str = json.dumps(value, ensure_ascii=False)
    else:
        value_str = str(value)

    cell = ws.cell(row=row_num, column=col_idx, value=value_str)
    cell.alignment = alignment
    return cell


def write_multi_select_answer(ws, row_num, answer, col_map, alignment):
    """
    Handle writing multi-select answers to individual option columns.

    For multi-select questions, each option gets its own column with a marker
    (True) to indicate if that option was selected.

    Args:
        ws: Worksheet object
        row_num: Row number to write to
        answer: The answer value (expected to be list, dict with 'answer', or other)
        col_map: Dict with '_options' key mapping option names to column indices
        alignment: Alignment object for cells
    """
    options_map = col_map.get("_options", {})

    # Extract the actual list of selected options
    selected_options = []

    if isinstance(answer, list):
        # Direct list of selected options
        selected_options = answer
    elif isinstance(answer, dict):
        if "answer" in answer:
            # Comment+answer structure
            inner_answer = answer.get("answer")
            if isinstance(inner_answer, list):
                selected_options = inner_answer
            elif inner_answer:
                selected_options = [inner_answer]
        elif "other" in answer:
            # Other option with custom text
            selected_options = ["Other"]
    elif answer:
        # Single value
        selected_options = [answer]

    # Write markers for each option
    for option, col_idx in options_map.items():
        # Check if this option was selected
        is_selected = False
        other_text = None

        for selected in selected_options:
            # Handle both direct option match and dict-based "other" answers
            if isinstance(selected, dict) and "other" in selected:
                if option.lower().startswith("other"):
                    is_selected = True
                    other_text = selected.get("other", "")
                    break
            elif str(selected) == str(option):
                is_selected = True
                break
            # Handle "Other: custom text" format - match against "Other" option
            elif (
                isinstance(selected, str)
                and selected.startswith("Other:")
                and option.lower() == "other"
            ):
                is_selected = True
                # Extract the custom text after "Other: "
                other_text = selected[6:].strip()  # Remove "Other:" prefix
                break

        # Write True for selected, empty for not selected
        # For "Other" options, include the custom text
        cell = ws.cell(row=row_num, column=col_idx)
        if is_selected:
            if other_text:
                cell.value = f"True: {other_text}"
            else:
                cell.value = "True"
        else:
            cell.value = ""
        cell.alignment = alignment


def write_complex_answer(ws, row_num, answer, col_map, alignment):
    """
    Handle writing complex dict answers to multiple sub-columns.

    This function handles different types of complex answers:
    - Regular dict: Spread values across sub-columns
    - Comment+answer dict: Handle nested structure
    - Non-dict in dict columns: Put value in first column

    Args:
        ws: Worksheet object
        row_num: Row number to write to
        answer: The answer value (expected to be dict, but handles other types)
        col_map: Dict mapping subfield names to column indices
        alignment: Alignment object for cells
    """
    if not isinstance(answer, dict):
        # Answer is not a dict but question has sub-columns
        # Put entire answer in first sub-column
        first_col_idx = list(col_map.values())[0]
        write_answer_to_cell(ws, row_num, first_col_idx, answer, alignment)
        # Fill other columns with empty
        for i, col_idx in enumerate(col_map.values()):
            if i > 0:
                ws.cell(row=row_num, column=col_idx, value="")
    elif "answer" in answer and "comment" in answer:
        # Comment+answer structure (e.g., from comment box questions)
        inner_answer = answer.get("answer")
        comment = answer.get("comment", "")

        if isinstance(inner_answer, dict):
            for subfield, col_idx in col_map.items():
                value = comment if subfield == "comment" else inner_answer.get(subfield, "")
                write_answer_to_cell(ws, row_num, col_idx, value if value else "", alignment)
        else:
            for subfield, col_idx in col_map.items():
                if subfield == "answer":
                    value = inner_answer
                elif subfield == "comment":
                    value = comment
                else:
                    value = ""
                write_answer_to_cell(ws, row_num, col_idx, value if value else "", alignment)
    else:
        # Regular dict answer (e.g., form fields, grid responses)
        for subfield, col_idx in col_map.items():
            value = answer.get(subfield, "")
            write_answer_to_cell(ws, row_num, col_idx, value if value else "", alignment)


def write_data_rows(
    ws,
    sessions,
    all_questions,
    question_column_map,
    is_completed_col,
    last_activity_col,
    styles,
    question_subfields,
    multi_select_questions,
):
    """
    Write all data rows to the worksheet.

    Args:
        ws: Worksheet object
        sessions: Dict of session data
        all_questions: List of Question objects
        question_column_map: Dict mapping question IDs to column indices
        is_completed_col: Column index for "Is Completed"
        last_activity_col: Column index for "Last Activity"
        styles: Dict of style objects
        question_subfields: Dict mapping question IDs to subfield names
        multi_select_questions: Dict mapping question IDs to list of options
    """
    # Determine starting row based on whether we have multi-select questions or subfields
    # Start after two header rows for consistency across all question types
    row_num = 3
    session_num = 0
    alignment = Alignment(wrap_text=True, vertical="top")

    for session_id, session_data in sessions.items():
        session_num += 1
        ws.cell(row=row_num, column=1, value=session_num)

        # Is Completed status with color coding (column 2)
        is_completed_cell = ws.cell(row=row_num, column=2)
        is_completed_cell.value = "Yes" if session_data["is_completed"] else "No"
        is_completed_cell.fill = (
            styles["completed_fill"] if session_data["is_completed"] else styles["incomplete_fill"]
        )

        # Write Respondent ID (last segment of survey_response_id UUID for completed, or session_id for partial) (column 3)
        respondent_id = ""
        if session_data.get("is_completed") and session_data.get("survey_response_id"):
            # For completed responses: use survey_response_id (already formatted as last part after split)
            respondent_id = session_data["survey_response_id"]
        elif not session_data.get("is_completed") and session_data.get("session_id"):
            # For partial responses: extract last segment of session_id (if UUID format) or last 8 characters
            session_id_str = str(session_data["session_id"])
            if "-" in session_id_str:
                respondent_id = session_id_str.split("-")[-1]
            else:
                # If not UUID format, use last 8 characters
                respondent_id = session_id_str[-8:] if len(session_id_str) >= 8 else session_id_str
        # Capitalize the respondent ID value (not the column header)
        ws.cell(row=row_num, column=3, value=respondent_id.upper() if respondent_id else "")

        # Fill in answers for each question
        for question in all_questions:
            answer = session_data["questions"].get(question.id)
            col_map = question_column_map[question.id]

            if answer is not None:
                if "_main" in col_map:
                    # Simple question - single column
                    write_answer_to_cell(ws, row_num, col_map["_main"], answer, alignment)
                elif "_type" in col_map and col_map["_type"] == "multi_select":
                    # Multi-select question - spread across option columns
                    write_multi_select_answer(ws, row_num, answer, col_map, alignment)
                else:
                    # Complex question - multiple sub-columns
                    write_complex_answer(ws, row_num, answer, col_map, alignment)
            else:
                # Empty answer - fill with empty strings
                if "_main" in col_map:
                    ws.cell(row=row_num, column=col_map["_main"], value="")
                elif "_type" in col_map and col_map["_type"] == "multi_select":
                    # Empty multi-select - fill all option columns
                    for col_idx in col_map.get("_options", {}).values():
                        ws.cell(row=row_num, column=col_idx, value="")
                else:
                    for col_idx in col_map.values():
                        ws.cell(row=row_num, column=col_idx, value="")

        # Last Activity timestamp
        last_activity_cell = ws.cell(row=row_num, column=last_activity_col)
        last_activity_cell.value = session_data["last_activity"].strftime("%Y-%m-%d %H:%M:%S")

        row_num += 1


def format_worksheet(
    ws, last_activity_col, all_questions, question_subfields, multi_select_questions
):
    """
    Apply formatting to the worksheet (column widths, row heights, freeze panes).

    Args:
        ws: Worksheet object
        last_activity_col: Index of the last column with data
        all_questions: List of Question objects
        question_subfields: Dict mapping question IDs to subfield names
        multi_select_questions: Dict mapping question IDs to list of options
    """
    # Column widths
    ws.column_dimensions["A"].width = 8  # Number column (narrow)
    ws.column_dimensions["B"].width = 15  # Is Completed column
    ws.column_dimensions["C"].width = 15  # Respondent ID column
    for col_idx in range(4, last_activity_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25

    # Row heights for headers
    # Always set two header rows (row 1 and 2) and freeze below them
    ws.row_dimensions[1].height = 30  # Main header row
    ws.row_dimensions[2].height = 25  # Second header row
    freeze_row = "A3"  # Freeze first 2 rows consistently

    # Freeze panes based on header structure
    ws.freeze_panes = freeze_row


def create_worksheet_with_data(
    wb, sheet_name, sessions, all_questions, question_subfields, multi_select_questions, styles
):
    """
    Create a complete worksheet with headers, data, and formatting.

    This is the main orchestration function that combines all worksheet building steps.

    Args:
        wb: Workbook object
        sheet_name: Name for the worksheet
        sessions: Dict of session data to include
        all_questions: List of Question objects
        question_subfields: Dict mapping question IDs to subfield names
        multi_select_questions: Dict mapping question IDs to list of options
        styles: Dict of style objects

    Returns:
        Worksheet: The created and populated worksheet object
    """
    ws = wb.create_sheet(title=sheet_name)

    # Build headers
    question_column_map, is_completed_col, last_activity_col = build_excel_headers(
        ws, all_questions, question_subfields, multi_select_questions, styles
    )

    # Write data rows
    write_data_rows(
        ws,
        sessions,
        all_questions,
        question_column_map,
        is_completed_col,
        last_activity_col,
        styles,
        question_subfields,
        multi_select_questions,
    )

    # Format worksheet
    format_worksheet(
        ws, last_activity_col, all_questions, question_subfields, multi_select_questions
    )

    return ws


def create_empty_sheet(wb, sheet_name, message="No data found"):
    """
    Create an empty worksheet with a message.

    Args:
        wb: Workbook object
        sheet_name: Name for the worksheet
        message: Message to display in the sheet

    Returns:
        Worksheet: The created worksheet
    """
    ws = wb.create_sheet(title=sheet_name)
    styles = create_excel_styles()

    cell = ws.cell(row=1, column=1, value=message)
    cell.fill = styles["header_fill"]
    cell.font = styles["header_font"]
    cell.alignment = styles["header_alignment"]
    ws.column_dimensions["A"].width = 40

    return ws
