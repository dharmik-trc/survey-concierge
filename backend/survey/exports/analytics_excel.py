"""
Analytics Excel Export Utilities

This module handles exporting analytics results to Excel format.
"""

from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .excel_styles import create_excel_styles


def create_analytics_excel(
    analytics: Dict, questions: List, survey_title: str, all_questions: List = None
) -> openpyxl.Workbook:
    """
    Create an Excel workbook with analytics results.

    Args:
        analytics: Dict mapping question_id to analytics data
        questions: List of Question objects (ordered) - may be filtered
        survey_title: Title of the survey
        all_questions: Optional full list of all questions (ordered) - used for correct numbering

    Returns:
        openpyxl.Workbook object
    """
    wb = openpyxl.Workbook()
    styles = create_excel_styles()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create main analytics sheet
    ws = wb.create_sheet(title="Analytics")

    # Style definitions
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Create mapping from question_id to position in all_questions list (for correct numbering)
    # Use all_questions if provided, otherwise use questions (assumes questions is the full list)
    questions_for_numbering = all_questions if all_questions is not None else questions
    question_number_map = {}
    for idx, q in enumerate(questions_for_numbering, start=1):
        question_number_map[q.id] = idx

    row_num = 1

    for question in questions:
        question_id = question.id
        if question_id not in analytics:
            continue

        # Use position in original all_questions list to match data export numbering
        # This maintains original question numbers even when some questions are filtered out
        question_num = question_number_map.get(question_id, (question.order or 0) + 1)
        analytics_data = analytics[question_id]

        # Question header
        ws.merge_cells(f"A{row_num}:D{row_num}")
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"Q{question_num}: {question.question_text}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

        row_num += 1

        # Write analytics based on type first
        question_type = analytics_data.get("type", "unknown")

        if question_type == "choice":
            row_num = write_choice_analytics(ws, analytics_data, row_num, styles, border)
        elif question_type == "grid":
            row_num = write_grid_analytics(ws, analytics_data, row_num, styles, border)
        elif question_type == "form_fields_numeric":
            row_num = write_form_fields_numeric_analytics(
                ws, analytics_data, row_num, styles, border
            )
        elif question_type == "numeric":
            # Explicitly call and verify numeric analytics are written
            initial_row = row_num
            row_num = write_numeric_analytics(ws, analytics_data, row_num, styles, border)
            # Verify rows were written (row_num should be at least initial_row + 9: 1 header + 8 stats)
            if row_num <= initial_row:
                # Fallback: write stats manually if function didn't write anything
                row_num = initial_row
                ws.cell(row=row_num, column=1, value="Statistic").border = border
                ws.cell(row=row_num, column=2, value="Value").border = border
                row_num += 1
                stats = [
                    ("Count", analytics_data.get("count", 0)),
                    ("Min", analytics_data.get("min", "N/A")),
                    ("Q1", analytics_data.get("q1", "N/A")),
                    ("Median", analytics_data.get("median", "N/A")),
                    ("Q3", analytics_data.get("q3", "N/A")),
                    ("Max", analytics_data.get("max", "N/A")),
                    ("Average", analytics_data.get("average", "N/A")),
                    ("Sum", analytics_data.get("sum", "N/A")),
                ]
                for stat_name, stat_value in stats:
                    ws.cell(row=row_num, column=1, value=str(stat_name)).border = border
                    ws.cell(row=row_num, column=2, value=str(stat_value)).border = border
                    row_num += 1
        else:
            # Other type or error message
            cell = ws.cell(row=row_num, column=1)
            cell.value = analytics_data.get("message", "No analytics available")
            cell.font = Font(italic=True)
            row_num += 1

        # Write answered/skipped summary row (after analytics data) - separate columns
        answered_count = analytics_data.get("answered_count", 0)
        skipped_count = analytics_data.get("skipped_count", 0)

        row_num += 1  # Add spacing
        ws.cell(row=row_num, column=1, value="Answered").border = border
        ws.cell(row=row_num, column=2, value="").border = border
        ws.cell(row=row_num, column=3, value=answered_count).border = border
        row_num += 1
        ws.cell(row=row_num, column=1, value="Skipped").border = border
        ws.cell(row=row_num, column=2, value="").border = border
        ws.cell(row=row_num, column=3, value=skipped_count).border = border
        row_num += 1

        # Write comments table if there are comments
        comments = analytics_data.get("comments", [])
        if comments:
            row_num = write_comments_table(ws, comments, row_num, styles, border)

        # Add spacing row
        row_num += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 40  # Answer Choices / Respondent ID / Answered/Skipped
    ws.column_dimensions["B"].width = 15  # Percentage
    ws.column_dimensions["C"].width = 12  # Count / Response Date
    ws.column_dimensions["D"].width = 50  # Comments
    ws.column_dimensions["E"].width = 15  # Tags

    # Freeze first row
    ws.freeze_panes = "A2"

    return wb


def write_choice_analytics(
    ws, analytics_data: Dict, start_row: int, styles: Dict, border: Border
) -> int:
    """Write choice question analytics to worksheet."""
    row_num = start_row

    # Headers - separate columns: Answer Choices, Percentage, Count
    headers = ["Answer Choices", "Percentage", "Count"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row_num += 1

    # Data rows - separate columns for Percentage and Count
    results = analytics_data.get("results", [])

    for result in results:
        option = result["option"]
        count = result["count"]
        percentage = result["percentage"]

        ws.cell(row=row_num, column=1, value=option).border = border
        ws.cell(row=row_num, column=2, value=f"{percentage}%").border = border
        ws.cell(row=row_num, column=3, value=count).border = border

        row_num += 1

    return row_num


def write_grid_analytics(
    ws, analytics_data: Dict, start_row: int, styles: Dict, border: Border
) -> int:
    """Write grid question analytics to worksheet."""
    row_num = start_row

    rows_data = analytics_data.get("rows", {})

    if not rows_data:
        cell = ws.cell(row=row_num, column=1)
        cell.value = analytics_data.get("message", "No data available")
        return row_num + 1

    # For each row in the grid
    for row_name, row_stats in rows_data.items():
        # Row header
        ws.merge_cells(f"A{row_num}:D{row_num}")
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"Row: {row_name} (Total Responses: {row_stats['total_responses']})"
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

        row_num += 1

        # Column headers - match format: Column, Percentage, Count (same as choice questions)
        headers = ["Column", "Percentage", "Count"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        row_num += 1

        # Column data - Percentage first, then Count (matches choice questions format)
        for col_stat in row_stats["columns"]:
            ws.cell(row=row_num, column=1, value=col_stat["column"]).border = border
            ws.cell(row=row_num, column=2, value=f"{col_stat['percentage']}%").border = border
            ws.cell(row=row_num, column=3, value=col_stat["count"]).border = border
            row_num += 1

        # Spacing between rows
        row_num += 1

    return row_num


def write_comments_table(
    ws, comments: List[Dict], start_row: int, styles: Dict, border: Border
) -> int:
    """Write comments table with Respondent ID, Response Date, and Comments."""
    row_num = start_row

    # Comments header
    ws.merge_cells(f"A{row_num}:D{row_num}")
    cell = ws.cell(row=row_num, column=1)
    cell.value = f"Any comments? ({len(comments)} responses)"
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

    row_num += 1

    # Table headers
    headers = ["Respondent ID", "Response Date", "Any comments?", "Tags"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row_num += 1

    # Comments data rows
    for comment_data in comments:
        response_id = comment_data.get("response_id", "")
        submitted_at = comment_data.get("submitted_at")
        comment_text = comment_data.get("comment", "")

        # Format date
        date_str = ""
        if submitted_at:
            try:
                if hasattr(submitted_at, "strftime"):
                    date_str = submitted_at.strftime("%b %d %Y %I:%M %p")
                else:
                    date_str = str(submitted_at)
            except (AttributeError, TypeError):
                date_str = str(submitted_at)

        ws.cell(row=row_num, column=1, value=response_id).border = border
        ws.cell(row=row_num, column=2, value=date_str).border = border
        ws.cell(row=row_num, column=3, value=comment_text).border = border
        ws.cell(row=row_num, column=4, value="").border = border

        row_num += 1

    return row_num


def write_numeric_analytics(
    ws, analytics_data: Dict, start_row: int, styles: Dict, border: Border
) -> int:
    """Write numeric question analytics to worksheet.

    Always displays the statistics table with:
    - Count, Min, Q1, Median, Q3, Max, Average, Sum
    """
    try:
        row_num = start_row

        # Always write the headers first - MUST write these
        header_cell1 = ws.cell(row=row_num, column=1)
        header_cell1.value = "Statistic"
        header_cell1.font = Font(bold=True, size=11)
        header_cell1.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_cell1.alignment = Alignment(horizontal="center", vertical="center")
        header_cell1.border = border

        header_cell2 = ws.cell(row=row_num, column=2)
        header_cell2.value = "Value"
        header_cell2.font = Font(bold=True, size=11)
        header_cell2.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_cell2.alignment = Alignment(horizontal="center", vertical="center")
        header_cell2.border = border

        row_num += 1

        # Get count - default to 0 if not present
        count = analytics_data.get("count", 0)
        try:
            count = int(count) if count is not None else 0
        except (ValueError, TypeError):
            count = 0

        # Check if we have valid numeric data (count > 0 and min exists)
        has_valid_data = count > 0 and analytics_data.get("min") is not None

        # Helper to safely get stat values
        def get_stat_value(key):
            if has_valid_data:
                val = analytics_data.get(key)
                return val if val is not None else "N/A"
            return "N/A"

        # ALWAYS write all 8 statistics - no conditions, no early returns
        stats = [
            ("Count", count),
            ("Min", get_stat_value("min")),
            ("Q1", get_stat_value("q1")),
            ("Median", get_stat_value("median")),
            ("Q3", get_stat_value("q3")),
            ("Max", get_stat_value("max")),
            ("Average", get_stat_value("average")),
            ("Sum", get_stat_value("sum")),
        ]

        # Write each statistic row - EXECUTE THIS LOOP ALWAYS
        for stat_name, stat_value in stats:
            # Write statistic name
            cell1 = ws.cell(row=row_num, column=1)
            cell1.value = str(stat_name)
            cell1.border = border
            cell1.font = Font(size=10)
            cell1.alignment = Alignment(horizontal="left", vertical="center")

            # Write statistic value
            cell2 = ws.cell(row=row_num, column=2)
            # Convert value to string, handle None
            value_str = str(stat_value) if stat_value is not None else "N/A"
            cell2.value = value_str
            cell2.border = border
            cell2.font = Font(size=10)
            cell2.alignment = Alignment(horizontal="right", vertical="center")

            row_num += 1

        # Return the next row number after all statistics (should be start_row + 9: 1 header + 8 stats)
        return row_num

    except Exception as e:
        # If anything fails, at least write a message and return next row
        import traceback

        print(f"ERROR in write_numeric_analytics: {e}")
        print(traceback.format_exc())
        # Write error message to Excel
        error_cell = ws.cell(row=start_row, column=1)
        error_cell.value = f"Error writing numeric analytics: {str(e)}"
        error_cell.font = Font(color="FF0000", italic=True)
        return start_row + 1


def write_form_fields_numeric_analytics(
    ws, analytics_data: Dict, start_row: int, styles: Dict, border: Border
) -> int:
    """Write form_fields numeric analytics to worksheet.

    Displays statistics for each numeric subfield separately.
    """
    row_num = start_row
    subfield_analytics = analytics_data.get("subfields", {})

    if not subfield_analytics:
        # No numeric subfields
        cell = ws.cell(row=row_num, column=1)
        cell.value = "No numeric subfields found"
        cell.font = Font(italic=True)
        return row_num + 1

    # Write analytics for each numeric subfield
    for subfield_name, subfield_stats in subfield_analytics.items():
        # Subfield header
        ws.merge_cells(f"A{row_num}:B{row_num}")
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"Subfield: {subfield_name}"
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        row_num += 1

        # Headers
        header_cell1 = ws.cell(row=row_num, column=1)
        header_cell1.value = "Statistic"
        header_cell1.font = Font(bold=True, size=11)
        header_cell1.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_cell1.alignment = Alignment(horizontal="center", vertical="center")
        header_cell1.border = border

        header_cell2 = ws.cell(row=row_num, column=2)
        header_cell2.value = "Value"
        header_cell2.font = Font(bold=True, size=11)
        header_cell2.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_cell2.alignment = Alignment(horizontal="center", vertical="center")
        header_cell2.border = border
        row_num += 1

        # Write statistics for this subfield
        stats = [
            ("Count", subfield_stats.get("count", 0)),
            ("Min", subfield_stats.get("min", "N/A")),
            ("Q1", subfield_stats.get("q1", "N/A")),
            ("Median", subfield_stats.get("median", "N/A")),
            ("Q3", subfield_stats.get("q3", "N/A")),
            ("Max", subfield_stats.get("max", "N/A")),
            ("Average", subfield_stats.get("average", "N/A")),
            ("Sum", subfield_stats.get("sum", "N/A")),
        ]

        for stat_name, stat_value in stats:
            cell1 = ws.cell(row=row_num, column=1)
            cell1.value = str(stat_name)
            cell1.border = border
            cell1.font = Font(size=10)
            cell1.alignment = Alignment(horizontal="left", vertical="center")

            cell2 = ws.cell(row=row_num, column=2)
            cell2.value = str(stat_value) if stat_value is not None else "N/A"
            cell2.border = border
            cell2.font = Font(size=10)
            cell2.alignment = Alignment(horizontal="right", vertical="center")
            row_num += 1

        # Add spacing between subfields
        row_num += 1

    return row_num
