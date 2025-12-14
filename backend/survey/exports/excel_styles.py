"""
Excel Styling Utilities

This module contains all styling-related functions and constants for Excel exports.
"""

from openpyxl.styles import Alignment, Font, PatternFill


def create_excel_styles():
    """
    Create and return all Excel styling objects.

    Returns:
        dict: Dictionary of style objects with the following keys:
            - header_fill: Dark indigo fill for main headers
            - subheader_fill: Light indigo fill for sub-headers
            - header_font: Bold white font for headers
            - subheader_font: Bold white font for sub-headers (smaller)
            - header_alignment: Center-aligned with text wrapping
            - completed_fill: Light green fill for completed responses
            - incomplete_fill: Light yellow fill for incomplete responses
    """
    return {
        "header_fill": PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid"),
        "subheader_fill": PatternFill(start_color="818CF8", end_color="818CF8", fill_type="solid"),
        "header_font": Font(bold=True, color="FFFFFF", size=12),
        "subheader_font": Font(bold=True, color="FFFFFF", size=10),
        "header_alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "completed_fill": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "incomplete_fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }


def get_cell_alignment(wrap_text=True, vertical="top", horizontal="left"):
    """
    Create a cell alignment object with commonly used settings.

    Args:
        wrap_text: Whether to wrap text in cells (default: True)
        vertical: Vertical alignment (default: "top")
        horizontal: Horizontal alignment (default: "left")

    Returns:
        Alignment: OpenPyXL Alignment object
    """
    return Alignment(wrap_text=wrap_text, vertical=vertical, horizontal=horizontal)
